#!/usr/bin/env python3
"""
Create sample Excel data for testing the Excel to PPT generator.
"""

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Headers
headers = [
    'Application Name',
    'Application Hosting Country',
    'LOBT',
    'Application Brief Description',
    'DATA DOMAIN FULL',
    'DOMAIN CATEGORY',
    'DOMAIN NAME',
    'SUB-DATA DOMAIN',
    'SUB-SUB-DATA DOMAIN',
    'ACTIVITY',
    'MAPPING CONFIDENCE',
    'Host Application Code',
    'Platform',
    'Application Category',
    'Application Owner (Primary)',
    'Application Owner (Secondary)',
    'Application Manager(Primary)',
    'Application Manager(Secondary)'
]

for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# Sample data
sample_data = [
    # Customer Data Domain
    ['Customer Portal', 'US', 'Retail Banking', 'Main customer interface', 'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Personal Information', 'Data Collection', 'High', 'APP001', 'Web', 'Frontend', 'John Smith', 'Jane Doe', 'Bob Wilson', 'Alice Brown'],
    ['CRM System', 'UK', 'Retail Banking', 'Customer relationship management', 'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Contact Details', 'Data Update', 'High', 'APP002', 'Enterprise', 'Backend', 'John Smith', 'Jane Doe', 'Bob Wilson', 'Alice Brown'],
    ['KYC Application', 'US', 'Compliance', 'Know your customer verification', 'Customer Data', 'Master Data', 'Customer', 'Customer Verification', 'KYC Documents', 'Verification', 'High', 'APP003', 'Web', 'Backend', 'Mary Johnson', 'Tom Lee', 'Bob Wilson', 'Alice Brown'],
    ['Mobile Banking', 'SG', 'Digital Banking', 'Mobile banking application', 'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Personal Information', 'Data View', 'Medium', 'APP004', 'Mobile', 'Frontend', 'John Smith', 'Jane Doe', 'Chris Davis', 'Eve Miller'],
    
    # Transaction Data Domain
    ['Payment Gateway', 'US', 'Payments', 'Process payments', 'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Real-time Payments', 'Processing', 'High', 'APP010', 'Enterprise', 'Backend', 'Sarah Connor', 'Mike Ross', 'David Chen', 'Lisa Wang'],
    ['Wire Transfer System', 'UK', 'Treasury', 'Wire transfer processing', 'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Wire Transfers', 'Processing', 'High', 'APP011', 'Mainframe', 'Backend', 'Sarah Connor', 'Mike Ross', 'David Chen', 'Lisa Wang'],
    ['ACH Processor', 'US', 'Payments', 'ACH batch processing', 'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'ACH Payments', 'Batch Processing', 'High', 'APP012', 'Mainframe', 'Backend', 'Sarah Connor', 'Mike Ross', 'David Chen', 'Lisa Wang'],
    ['Transaction History', 'HK', 'Digital Banking', 'View transaction history', 'Transaction Data', 'Operational', 'Payments', 'Transaction Records', 'Historical Data', 'Data View', 'Medium', 'APP013', 'Web', 'Frontend', 'John Smith', 'Jane Doe', 'Chris Davis', 'Eve Miller'],
    
    # Risk Data Domain
    ['Credit Scoring', 'US', 'Risk Management', 'Credit score calculation', 'Risk Data', 'Analytics', 'Risk', 'Credit Risk', 'Scoring Models', 'Calculation', 'High', 'APP020', 'Enterprise', 'Backend', 'Risk Admin', 'Risk Team', 'Risk Manager', 'Compliance Lead'],
    ['Fraud Detection', 'UK', 'Risk Management', 'Real-time fraud detection', 'Risk Data', 'Analytics', 'Risk', 'Fraud Risk', 'Transaction Monitoring', 'Real-time Analysis', 'High', 'APP021', 'Enterprise', 'Backend', 'Risk Admin', 'Risk Team', 'Risk Manager', 'Compliance Lead'],
    ['AML Screening', 'SG', 'Compliance', 'Anti-money laundering checks', 'Risk Data', 'Analytics', 'Risk', 'AML Risk', 'Screening', 'Compliance Check', 'High', 'APP022', 'Enterprise', 'Backend', 'Compliance Admin', 'Compliance Team', 'Risk Manager', 'Compliance Lead'],
    ['Market Risk System', 'HK', 'Risk Management', 'Market risk calculation', 'Risk Data', 'Analytics', 'Risk', 'Market Risk', 'VaR Calculation', 'Calculation', 'High', 'APP023', 'Enterprise', 'Backend', 'Risk Admin', 'Risk Team', 'Risk Manager', 'Compliance Lead'],
    
    # Regulatory Data Domain
    ['Regulatory Reporting', 'US', 'Compliance', 'Generate regulatory reports', 'Regulatory Data', 'Compliance', 'Regulatory', 'Regulatory Reports', 'Fed Reports', 'Report Generation', 'High', 'APP030', 'Enterprise', 'Backend', 'Reg Admin', 'Reg Team', 'Compliance Manager', 'Audit Lead'],
    ['Basel III Calculator', 'UK', 'Risk Management', 'Basel capital calculation', 'Regulatory Data', 'Compliance', 'Regulatory', 'Capital Requirements', 'Basel Calculations', 'Calculation', 'High', 'APP031', 'Enterprise', 'Backend', 'Reg Admin', 'Reg Team', 'Compliance Manager', 'Audit Lead'],
    ['MiFID Reporting', 'UK', 'Compliance', 'MiFID II compliance', 'Regulatory Data', 'Compliance', 'Regulatory', 'Trade Reporting', 'MiFID Reports', 'Report Generation', 'High', 'APP032', 'Enterprise', 'Backend', 'Reg Admin', 'Reg Team', 'Compliance Manager', 'Audit Lead'],
]

for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column].width = min(max_length + 2, 30)

wb.save('sample_data.xlsx')
print("Created sample_data.xlsx with test data")
