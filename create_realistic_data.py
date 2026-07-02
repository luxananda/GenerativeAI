#!/usr/bin/env python3
"""
Create realistic banking application data for testing.
Simulates a real bank's application portfolio mapped to data domains.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Application Data Mapping"

# Headers
headers = [
    'Application Name',
    'Application Hosting Country',
    'LOBT',
    'Application Brief Description',
    'DATA DOMAIN FULL',
    'DOMAIN CATEGORY',
    'DOMAIN NAME',
    'SUB-DATA DOMAIN',
    'SUB-SUB-DATA DOMAIN',
    'ACTIVITY',
    'MAPPING CONFIDENCE',
    'Host Application Code',
    'Platform',
    'Application Category',
    'Application Owner (Primary)',
    'Application Owner (Secondary)',
    'Application Manager(Primary)',
    'Application Manager(Secondary)'
]

# Style headers
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Realistic banking application data
# Structure: [App Name, Country, LOBT, Description, Data Domain Full, Domain Category, Domain Name, 
#             Sub-Data Domain, Sub-Sub-Data Domain, Activity, Confidence, App Code, Platform, 
#             App Category, Owner1, Owner2, Manager1, Manager2]

data = [
    # =========================================================================
    # CUSTOMER DATA DOMAIN
    # =========================================================================
    # Customer Demographics - Personal Information
    ['Customer 360 Portal', 'US', 'Retail Banking', 'Unified customer view platform', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Personal Information', 
     'Data Collection', 'High', 'CUS001', 'Cloud - AWS', 'Customer Facing', 
     'Sarah Mitchell', 'David Park', 'Jennifer Adams', 'Michael Brown'],
    
    ['Mobile Banking App', 'US', 'Digital Banking', 'Consumer mobile banking application', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Personal Information', 
     'Data View', 'High', 'MOB001', 'Mobile - iOS/Android', 'Customer Facing', 
     'Sarah Mitchell', 'David Park', 'Jennifer Adams', 'Michael Brown'],
    
    ['Internet Banking Platform', 'UK', 'Digital Banking', 'Web-based banking portal', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Personal Information', 
     'Data Update', 'High', 'WEB001', 'Web - Azure', 'Customer Facing', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    # Customer Demographics - Contact Details
    ['CRM Enterprise', 'US', 'Sales & Marketing', 'Customer relationship management', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Contact Details', 
     'Data Management', 'High', 'CRM001', 'Salesforce', 'Internal', 
     'Sarah Mitchell', 'David Park', 'Jennifer Adams', 'Michael Brown'],
    
    ['Contact Center System', 'IN', 'Operations', 'Call center management platform', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Demographics', 'Contact Details', 
     'Data Collection', 'Medium', 'CC001', 'On-Premise', 'Internal', 
     'Priya Sharma', 'Raj Patel', 'Anita Desai', 'Vikram Singh'],
    
    # Customer Verification - KYC
    ['KYC Onboarding', 'US', 'Compliance', 'Know Your Customer verification', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Verification', 'KYC Documents', 
     'Verification', 'High', 'KYC001', 'Cloud - AWS', 'Compliance', 
     'Mark Stevens', 'Linda Garcia', 'Thomas Wright', 'Nancy Kim'],
    
    ['Document Verification Engine', 'UK', 'Compliance', 'ID document verification AI', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Verification', 'KYC Documents', 
     'Document Processing', 'High', 'DVE001', 'Cloud - GCP', 'Compliance', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    ['Biometric Authentication', 'SG', 'Security', 'Facial recognition for authentication', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Verification', 'Biometric Data', 
     'Authentication', 'High', 'BIO001', 'Cloud - AWS', 'Security', 
     'Wei Chen', 'Michelle Tan', 'Kevin Lim', 'Grace Wong'],
    
    # Customer Preferences
    ['Preference Management', 'US', 'Digital Banking', 'Customer preferences and settings', 
     'Customer Data', 'Master Data', 'Customer', 'Customer Preferences', 'Communication Preferences', 
     'Data Management', 'Medium', 'PFM001', 'Cloud - Azure', 'Customer Facing', 
     'Sarah Mitchell', 'David Park', 'Jennifer Adams', 'Michael Brown'],
    
    # =========================================================================
    # TRANSACTION DATA DOMAIN
    # =========================================================================
    # Payment Processing - Real-time
    ['Real-time Payments Hub', 'US', 'Payments', 'RTP and instant payment processing', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Real-time Payments', 
     'Transaction Processing', 'High', 'RTP001', 'On-Premise - Mainframe', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    ['FedNow Gateway', 'US', 'Payments', 'Federal Reserve instant payments', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Real-time Payments', 
     'Transaction Processing', 'High', 'FED001', 'On-Premise', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    # Payment Processing - Wire Transfers
    ['SWIFT Gateway', 'UK', 'Treasury', 'International wire transfers', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Wire Transfers', 
     'Transaction Processing', 'High', 'SWT001', 'On-Premise', 'Treasury', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    ['Domestic Wire System', 'US', 'Treasury', 'Domestic wire transfer processing', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'Wire Transfers', 
     'Transaction Processing', 'High', 'DWS001', 'On-Premise - Mainframe', 'Treasury', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    # Payment Processing - ACH
    ['ACH Origination', 'US', 'Payments', 'ACH payment origination', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'ACH Payments', 
     'Batch Processing', 'High', 'ACH001', 'On-Premise - Mainframe', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    ['ACH Returns Handler', 'US', 'Payments', 'ACH return processing', 
     'Transaction Data', 'Operational', 'Payments', 'Payment Processing', 'ACH Payments', 
     'Exception Handling', 'High', 'ACH002', 'On-Premise', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    # Card Transactions
    ['Card Authorization Engine', 'US', 'Cards', 'Real-time card authorization', 
     'Transaction Data', 'Operational', 'Cards', 'Card Transactions', 'Authorization', 
     'Real-time Processing', 'High', 'CAE001', 'On-Premise', 'Cards', 
     'Patricia Moore', 'Christopher Davis', 'Angela White', 'Brian Taylor'],
    
    ['Merchant Settlement', 'US', 'Cards', 'Merchant payment settlement', 
     'Transaction Data', 'Operational', 'Cards', 'Card Transactions', 'Settlement', 
     'Batch Processing', 'High', 'MER001', 'On-Premise', 'Cards', 
     'Patricia Moore', 'Christopher Davis', 'Angela White', 'Brian Taylor'],
    
    ['Dispute Management', 'US', 'Cards', 'Chargeback and dispute handling', 
     'Transaction Data', 'Operational', 'Cards', 'Card Transactions', 'Disputes', 
     'Case Management', 'Medium', 'DIS001', 'Cloud - AWS', 'Cards', 
     'Patricia Moore', 'Christopher Davis', 'Angela White', 'Brian Taylor'],
    
    # Transaction Records
    ['Transaction Archive', 'US', 'Operations', 'Historical transaction storage', 
     'Transaction Data', 'Operational', 'Payments', 'Transaction Records', 'Historical Data', 
     'Data Storage', 'High', 'ARC001', 'Cloud - AWS S3', 'Data Platform', 
     'John Roberts', 'Mary Anderson', 'Steven Clark', 'Karen Lewis'],
    
    ['Statement Generator', 'US', 'Operations', 'Account statement generation', 
     'Transaction Data', 'Operational', 'Payments', 'Transaction Records', 'Statements', 
     'Report Generation', 'High', 'STM001', 'On-Premise', 'Operations', 
     'John Roberts', 'Mary Anderson', 'Steven Clark', 'Karen Lewis'],
    
    # =========================================================================
    # RISK DATA DOMAIN
    # =========================================================================
    # Credit Risk - Scoring
    ['Credit Decision Engine', 'US', 'Risk Management', 'Automated credit decisions', 
     'Risk Data', 'Analytics', 'Risk', 'Credit Risk', 'Scoring Models', 
     'Risk Calculation', 'High', 'CDE001', 'Cloud - AWS', 'Risk', 
     'Elizabeth Turner', 'William Harris', 'Jessica Martin', 'Andrew Jackson'],
    
    ['FICO Score Integration', 'US', 'Risk Management', 'Bureau score retrieval', 
     'Risk Data', 'Analytics', 'Risk', 'Credit Risk', 'Scoring Models', 
     'Data Integration', 'High', 'FIC001', 'API Gateway', 'Risk', 
     'Elizabeth Turner', 'William Harris', 'Jessica Martin', 'Andrew Jackson'],
    
    ['Portfolio Risk Monitor', 'US', 'Risk Management', 'Credit portfolio monitoring', 
     'Risk Data', 'Analytics', 'Risk', 'Credit Risk', 'Portfolio Analytics', 
     'Risk Monitoring', 'High', 'PRM001', 'Cloud - Azure', 'Risk', 
     'Elizabeth Turner', 'William Harris', 'Jessica Martin', 'Andrew Jackson'],
    
    # Fraud Risk
    ['Real-time Fraud Detection', 'US', 'Fraud', 'ML-based fraud detection', 
     'Risk Data', 'Analytics', 'Risk', 'Fraud Risk', 'Transaction Monitoring', 
     'Real-time Analysis', 'High', 'FRD001', 'Cloud - GCP', 'Fraud', 
     'Michelle Rodriguez', 'Jason Lee', 'Stephanie Brown', 'Ryan Miller'],
    
    ['Fraud Case Manager', 'US', 'Fraud', 'Fraud investigation workflow', 
     'Risk Data', 'Analytics', 'Risk', 'Fraud Risk', 'Case Management', 
     'Investigation', 'High', 'FCM001', 'Cloud - AWS', 'Fraud', 
     'Michelle Rodriguez', 'Jason Lee', 'Stephanie Brown', 'Ryan Miller'],
    
    ['Account Takeover Prevention', 'UK', 'Fraud', 'ATO detection and prevention', 
     'Risk Data', 'Analytics', 'Risk', 'Fraud Risk', 'Account Security', 
     'Prevention', 'High', 'ATO001', 'Cloud - AWS', 'Fraud', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    # AML Risk
    ['Transaction Monitoring AML', 'US', 'Compliance', 'AML transaction screening', 
     'Risk Data', 'Analytics', 'Risk', 'AML Risk', 'Transaction Screening', 
     'Compliance Monitoring', 'High', 'AML001', 'On-Premise', 'Compliance', 
     'Mark Stevens', 'Linda Garcia', 'Thomas Wright', 'Nancy Kim'],
    
    ['Sanctions Screening', 'US', 'Compliance', 'OFAC and sanctions list screening', 
     'Risk Data', 'Analytics', 'Risk', 'AML Risk', 'Sanctions', 
     'Screening', 'High', 'SAN001', 'On-Premise', 'Compliance', 
     'Mark Stevens', 'Linda Garcia', 'Thomas Wright', 'Nancy Kim'],
    
    ['SAR Filing System', 'US', 'Compliance', 'Suspicious activity reporting', 
     'Risk Data', 'Analytics', 'Risk', 'AML Risk', 'Regulatory Reporting', 
     'Report Filing', 'High', 'SAR001', 'On-Premise', 'Compliance', 
     'Mark Stevens', 'Linda Garcia', 'Thomas Wright', 'Nancy Kim'],
    
    # Market Risk
    ['VaR Calculator', 'UK', 'Treasury', 'Value at Risk calculation', 
     'Risk Data', 'Analytics', 'Risk', 'Market Risk', 'VaR Models', 
     'Risk Calculation', 'High', 'VAR001', 'On-Premise', 'Treasury', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    ['Market Data Feed', 'UK', 'Treasury', 'Real-time market data', 
     'Risk Data', 'Analytics', 'Risk', 'Market Risk', 'Market Data', 
     'Data Integration', 'High', 'MKT001', 'Cloud - Bloomberg', 'Treasury', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    # Operational Risk
    ['Incident Management', 'US', 'Operations', 'Operational incident tracking', 
     'Risk Data', 'Analytics', 'Risk', 'Operational Risk', 'Incident Tracking', 
     'Case Management', 'Medium', 'INC001', 'ServiceNow', 'Operations', 
     'John Roberts', 'Mary Anderson', 'Steven Clark', 'Karen Lewis'],
    
    ['Control Testing', 'US', 'Risk Management', 'Control effectiveness testing', 
     'Risk Data', 'Analytics', 'Risk', 'Operational Risk', 'Control Assessment', 
     'Testing', 'Medium', 'CTL001', 'Cloud - AWS', 'Risk', 
     'Elizabeth Turner', 'William Harris', 'Jessica Martin', 'Andrew Jackson'],
    
    # =========================================================================
    # REGULATORY DATA DOMAIN
    # =========================================================================
    # Regulatory Reports - Fed Reports
    ['FR Y-14 Reporting', 'US', 'Regulatory', 'Federal Reserve stress test reporting', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Regulatory Reports', 'Fed Reports', 
     'Report Generation', 'High', 'FRY001', 'On-Premise', 'Regulatory', 
     'Catherine Phillips', 'Edward Morgan', 'Nicole Baker', 'Scott Nelson'],
    
    ['Call Report System', 'US', 'Regulatory', 'FFIEC Call Report generation', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Regulatory Reports', 'Fed Reports', 
     'Report Filing', 'High', 'CAL001', 'On-Premise', 'Regulatory', 
     'Catherine Phillips', 'Edward Morgan', 'Nicole Baker', 'Scott Nelson'],
    
    # Capital Requirements
    ['Basel III Engine', 'UK', 'Risk Management', 'Basel capital calculations', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Capital Requirements', 'Basel Calculations', 
     'Risk Calculation', 'High', 'BAS001', 'On-Premise', 'Risk', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    ['RWA Calculator', 'US', 'Risk Management', 'Risk-weighted assets calculation', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Capital Requirements', 'RWA', 
     'Risk Calculation', 'High', 'RWA001', 'On-Premise', 'Risk', 
     'Elizabeth Turner', 'William Harris', 'Jessica Martin', 'Andrew Jackson'],
    
    ['Capital Planning Tool', 'US', 'Finance', 'Capital adequacy planning', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Capital Requirements', 'Capital Planning', 
     'Planning', 'High', 'CPT001', 'Cloud - Azure', 'Finance', 
     'Catherine Phillips', 'Edward Morgan', 'Nicole Baker', 'Scott Nelson'],
    
    # Trade Reporting
    ['MiFID II Reporter', 'UK', 'Regulatory', 'MiFID II transaction reporting', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Trade Reporting', 'MiFID Reports', 
     'Report Filing', 'High', 'MIF001', 'Cloud - AWS', 'Regulatory', 
     'James Wilson', 'Emma Thompson', 'Robert Clarke', 'Lisa Chen'],
    
    ['Dodd-Frank Reporting', 'US', 'Regulatory', 'Swap data repository reporting', 
     'Regulatory Data', 'Compliance', 'Regulatory', 'Trade Reporting', 'Dodd-Frank', 
     'Report Filing', 'High', 'DFR001', 'On-Premise', 'Regulatory', 
     'Catherine Phillips', 'Edward Morgan', 'Nicole Baker', 'Scott Nelson'],
    
    # =========================================================================
    # ACCOUNT DATA DOMAIN
    # =========================================================================
    # Account Management - Deposit Accounts
    ['Core Banking System', 'US', 'Retail Banking', 'Core deposit account system', 
     'Account Data', 'Master Data', 'Account', 'Account Management', 'Deposit Accounts', 
     'Account Processing', 'High', 'CBS001', 'On-Premise - Mainframe', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    ['Savings Account Manager', 'US', 'Retail Banking', 'Savings account management', 
     'Account Data', 'Master Data', 'Account', 'Account Management', 'Deposit Accounts', 
     'Interest Calculation', 'High', 'SAV001', 'On-Premise - Mainframe', 'Core Banking', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    # Account Management - Loan Accounts
    ['Loan Origination System', 'US', 'Lending', 'Consumer loan origination', 
     'Account Data', 'Master Data', 'Account', 'Account Management', 'Loan Accounts', 
     'Account Opening', 'High', 'LOS001', 'Cloud - AWS', 'Lending', 
     'Robert Johnson', 'Amy Williams', 'Charles Davis', 'Laura Thompson'],
    
    ['Mortgage Platform', 'US', 'Lending', 'Mortgage loan processing', 
     'Account Data', 'Master Data', 'Account', 'Account Management', 'Loan Accounts', 
     'Loan Processing', 'High', 'MTG001', 'Cloud - Azure', 'Lending', 
     'Robert Johnson', 'Amy Williams', 'Charles Davis', 'Laura Thompson'],
    
    ['Commercial Lending', 'US', 'Commercial Banking', 'Commercial loan management', 
     'Account Data', 'Master Data', 'Account', 'Account Management', 'Loan Accounts', 
     'Credit Management', 'High', 'CML001', 'On-Premise', 'Commercial', 
     'Peter Anderson', 'Rachel Green', 'Timothy Hall', 'Sandra Lee'],
    
    # Account Servicing
    ['Account Maintenance Hub', 'US', 'Operations', 'Account servicing operations', 
     'Account Data', 'Master Data', 'Account', 'Account Servicing', 'Maintenance', 
     'Account Update', 'High', 'AMH001', 'On-Premise', 'Operations', 
     'John Roberts', 'Mary Anderson', 'Steven Clark', 'Karen Lewis'],
    
    ['Overdraft Management', 'US', 'Operations', 'Overdraft protection system', 
     'Account Data', 'Master Data', 'Account', 'Account Servicing', 'Overdraft', 
     'Exception Processing', 'Medium', 'OVD001', 'On-Premise - Mainframe', 'Operations', 
     'John Roberts', 'Mary Anderson', 'Steven Clark', 'Karen Lewis'],
    
    # =========================================================================
    # PRODUCT DATA DOMAIN
    # =========================================================================
    # Product Catalog
    ['Product Catalog Manager', 'US', 'Product Management', 'Bank product definitions', 
     'Product Data', 'Reference', 'Product', 'Product Catalog', 'Product Definitions', 
     'Data Management', 'High', 'PCM001', 'Cloud - AWS', 'Product', 
     'Diana Ross', 'Frank Wilson', 'Helen Carter', 'Gary Mitchell'],
    
    ['Pricing Engine', 'US', 'Product Management', 'Product pricing calculations', 
     'Product Data', 'Reference', 'Product', 'Product Catalog', 'Pricing', 
     'Price Calculation', 'High', 'PRE001', 'Cloud - AWS', 'Product', 
     'Diana Ross', 'Frank Wilson', 'Helen Carter', 'Gary Mitchell'],
    
    ['Rate Management', 'US', 'Treasury', 'Interest rate management', 
     'Product Data', 'Reference', 'Product', 'Product Catalog', 'Interest Rates', 
     'Rate Setting', 'High', 'RTM001', 'On-Premise', 'Treasury', 
     'Richard Lee', 'Amanda Foster', 'Daniel Kim', 'Susan Martinez'],
    
    # Product Analytics
    ['Product Performance Analytics', 'US', 'Analytics', 'Product profitability analysis', 
     'Product Data', 'Reference', 'Product', 'Product Analytics', 'Performance Metrics', 
     'Analytics', 'Medium', 'PPA001', 'Cloud - Snowflake', 'Analytics', 
     'Diana Ross', 'Frank Wilson', 'Helen Carter', 'Gary Mitchell'],
]

# Write data
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')

# Adjust column widths
column_widths = {
    'A': 25,  # Application Name
    'B': 12,  # Country
    'C': 18,  # LOBT
    'D': 35,  # Description
    'E': 18,  # Data Domain Full
    'F': 15,  # Domain Category
    'G': 12,  # Domain Name
    'H': 22,  # Sub-Data Domain
    'I': 22,  # Sub-Sub-Data Domain
    'J': 20,  # Activity
    'K': 12,  # Confidence
    'L': 10,  # App Code
    'M': 20,  # Platform
    'N': 15,  # App Category
    'O': 18,  # Owner 1
    'P': 18,  # Owner 2
    'Q': 18,  # Manager 1
    'R': 18,  # Manager 2
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze header row
ws.freeze_panes = 'A2'

# Add autofilter
ws.auto_filter.ref = ws.dimensions

wb.save('bank_application_data.xlsx')
print(f"Created bank_application_data.xlsx with {len(data)} applications")
print("\nData Domains included:")
domains = set(row[4] for row in data)
for domain in sorted(domains):
    count = sum(1 for row in data if row[4] == domain)
    print(f"  - {domain}: {count} applications")
